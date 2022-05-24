
import openpyxl
import os


def generateTitleCase(input_string):  # Function to convert into title case
    # list of articles
    articles = ["a", "an", "the"]
    artiU = ["A", "AN", "THE"]
    artiM = ["An", "The"]
    arti = artiM + artiU
    # list of coordinating conjunctins
    conjunctions = ["and", "but", "for", "nor", "or", "so", "yet"]
    conjU = ["AND", "BUT", "FOR", "NOR", "OR", "SO", "YET"]
    conjM = ["And", "But", "For", "Nor", "Or", "So", "Yet"]
    conju = conjM + conjU
    # list of some short articles
    prepositions = ["in", "to", "for", "with", "on", "at", "from", "by", "about", "as", "into", "like", "through", "after", "over", "between", "out", "against", "during", "without", "before", "under", "around", "among", "of"]

    prepU = ["IN", "TO", "FOR", "WITH", "ON", "AT", "FROM", "BY", "ABOUT", "AS", "INTO", "LIKE", "THROUGH", "AFTER", "OVER", "BETWEEN", "OUT", "AGAINST", "DURING", "WITHOUT", "BEFORE", "UNDER", "AROUND", "AMONG", "OF"]
    prepM = ["In", "To", "For", "With", "On", "At", "From", "By", "About", "As", "Into", "Like", "Through", "After", "Over", "Between", "Out", "Against", "During", "Without", "Before", "Under", "Around", "Among", "Of", "oF"]
    prep = prepM + prepU
    # merging the 3 lists
    lower_case = articles + conjunctions + prepositions
    lowercase = conju + prep + arti

    # Main Work
    output_string = ""
    if input_string is not None:
        input_list = input_string.split(" ")
        for Cust in input_list:
            if Cust in lower_case:
                output_string += Cust + " "
            elif Cust in lowercase:
                output_string += Cust.lower() + " "
            else:
                temp = Cust.title()
                output_string += temp + " "
        return output_string


class Data:
    global Raw
    InstruList = []
    CharList = []
    Workbook = openpyxl.load_workbook("Files\\UserFile.xlsx")
    sheet = Workbook.active
    for i in range(2, sheet.max_row+1):
        Raw = sheet.cell(row=i, column=1).value
        print(Raw)
        input_text1 = Raw
        if input_text1 is not None:
            Instrument = generateTitleCase(input_text1).rstrip()
            InstruList.append(Instrument)
            print(Instrument)
            FirstChar = Instrument[0]
            CharList.append(FirstChar)
        else:
            Instrument = None
            InstruList.append(None)
            CharList.append(None)

    print(InstruList)
